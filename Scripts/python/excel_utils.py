#!/usr/bin/env python3
"""
Excel Utility Functions for Channel Director Automation

Provides helpers for creating Excel workbooks with consistent formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json


class ExcelReport:
    """Helper class for creating Excel reports."""
    
    def __init__(self, title="Report"):
        """Initialize Excel workbook with default styling."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Summary"
        self.title = title
        self.current_row = 1
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right_alignment = Alignment(horizontal="right", vertical="center")
    
    def add_sheet(self, sheet_name):
        """Add a new sheet to the workbook."""
        self.ws = self.wb.create_sheet(sheet_name)
        self.current_row = 1
        return self.ws
    
    def add_title(self, text, subtitle=None):
        """Add title to current sheet."""
        # Title
        self.ws.merge_cells(f'A{self.current_row}:F{self.current_row}')
        cell = self.ws[f'A{self.current_row}']
        cell.value = text
        cell.font = Font(bold=True, size=14)
        cell.alignment = self.center_alignment
        self.current_row += 1
        
        # Subtitle if provided
        if subtitle:
            self.ws.merge_cells(f'A{self.current_row}:F{self.current_row}')
            cell = self.ws[f'A{self.current_row}']
            cell.value = subtitle
            cell.font = Font(italic=True, size=10)
            cell.alignment = self.center_alignment
            self.current_row += 1
        
        # Blank row
        self.current_row += 1
    
    def add_header_row(self, headers, fill_color="366092", font_color="FFFFFF"):
        """Add header row with styling."""
        header_font = Font(bold=True, color=font_color, size=11)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        self.current_row += 1
    
    def add_data_row(self, values, bold=False, fill_color=None):
        """Add data row."""
        for col_num, value in enumerate(values, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = value
            cell.border = self.border
            
            if bold:
                cell.font = Font(bold=True)
            
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Alignment based on value type
            if isinstance(value, (int, float)):
                cell.alignment = self.right_alignment
            else:
                cell.alignment = self.left_alignment
        
        self.current_row += 1
    
    def add_table(self, headers, data, bold_first_col=False):
        """Add a complete table (headers + data)."""
        # Headers
        self.add_header_row(headers)
        
        # Data rows
        for row in data:
            values = [row.get(h) if isinstance(row, dict) else row[headers.index(h)] 
                     for h in headers]
            bold = bold_first_col  # Could make first column bold if needed
            self.add_data_row(values, bold=bold)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            for cell in self.ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add blank row
        self.current_row += 1
    
    def add_metric(self, label, value, target=None, status=None):
        """Add a key metric row."""
        self.ws[f'A{self.current_row}'] = label
        self.ws[f'B{self.current_row}'] = value
        
        if target:
            self.ws[f'C{self.current_row}'] = target
        
        if status:
            self.ws[f'D{self.current_row}'] = status
            # Color code: green for good, yellow for warning, red for bad
            if "↓" in str(status) or "Risk" in str(status):
                color = "FFC7CE"  # Light red
            elif "↑" in str(status) or "On Track" in str(status):
                color = "C6EFCE"  # Light green
            else:
                color = "FFFFEB"  # Light yellow
            self.ws[f'D{self.current_row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        self.current_row += 1
    
    def add_blank_rows(self, count=1):
        """Add blank rows."""
        self.current_row += count
    
    def add_metadata(self):
        """Add metadata footer."""
        self.add_blank_rows(1)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S UTC")
        self.ws[f'A{self.current_row}'] = f"Generated: {timestamp}"
        self.ws[f'A{self.current_row}'].font = Font(italic=True, size=9)
    
    def save(self, filepath):
        """Save workbook to file."""
        self.wb.save(filepath)
        print(f"✅ Saved Excel report: {filepath}")


def load_json_data(json_str):
    """Load and parse JSON data."""
    try:
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        print(f"❌ Invalid JSON: {e}")
        return None


def safe_get(obj, key, default=None):
    """Safely get value from dict or nested dict/path."""
    if isinstance(key, list):
        # Nested path like ["data", "revenue"]
        current = obj
        for k in key:
            if isinstance(current, dict):
                current = current.get(k, default)
            else:
                return default
        return current
    elif isinstance(obj, dict):
        return obj.get(key, default)
    return default


def format_currency(value, currency="EUR"):
    """Format number as currency."""
    if value is None:
        return "—"
    try:
        if currency == "EUR":
            return f"€{value:,.0f}"
        else:
            return f"${value:,.0f}"
    except:
        return str(value)


def format_percentage(value):
    """Format number as percentage."""
    if value is None:
        return "—"
    try:
        return f"{float(value):.1f}%"
    except:
        return str(value)


def get_status_emoji(status):
    """Get emoji based on status."""
    if status is None:
        return "—"
    status_str = str(status).lower()
    if "exceeding" in status_str or "excellent" in status_str or "strong" in status_str:
        return "✅"
    elif "on track" in status_str or "good" in status_str:
        return "✅"
    elif "risk" in status_str or "concern" in status_str:
        return "⚠️"
    elif "critical" in status_str or "danger" in status_str:
        return "🚨"
    else:
        return "—"


# ═══════════════════════════════════════════════════════════════════════════
# END OF EXCEL UTILS
# ═══════════════════════════════════════════════════════════════════════════
